#!/usr/bin/env python3
"""
importar_horarios.py — Parser de Excel de horarios en formato semanal (UPCT)

Lee ficheros del tipo:
    25-26_GIDI 1C.xlsx   (1er curso, ambos cuatrimestres)
    25-26_GIDI 2C.xlsx   (2º curso)
    ...

Cada fichero tiene hojas  1C_grupo1  y  2C_grupo1  con este esquema:
    Fila 0   "SEMANA N:  DD MES A DD MES"  en columnas de inicio de semana
    Fila 1   info (titulación, aula, descripción de grupo)
    Fila 2   días del mes (números)
    Fila 3   días de la semana (LUNES … VIERNES)
    Filas 4+ franjas y contenido de clases
    Col  3   etiqueta de franja ("9:00 - 10:50", "11:10 - 13:00", …)
    Col  4+  contenido de clases por semana × día

Contenido de celda (ejemplos):
    [524101003] Expresión Gráfica
    [524101003] Expresión Gráfica | INFO | Subgrupos: 2
    [524101001] Matemáticas I | LAB | Subgrupos: 1
    [524101011] Metodología del Diseño | Aula: PB2
    NO LECTIVO
"""

import io
import re

_SUBJECT_RE = re.compile(r'^\[(\d+)\]\s*(.+?)(?:\s*\|(.*))?$', re.DOTALL)
_SEMANA_RE  = re.compile(r'SEMANA\s+(\d+)', re.IGNORECASE)
_FRANJA_RE  = re.compile(r'^\d{1,2}:\d{2}')

# Normalización de días (con/sin tilde → con tilde, coherente con horarios.js y servidor)
DIAS_MAP = {
    'LUNES':     'LUNES',
    'MARTES':    'MARTES',
    'MIÉRCOLES': 'MIÉRCOLES',
    'MIERCOLES': 'MIÉRCOLES',
    'JUEVES':    'JUEVES',
    'VIERNES':   'VIERNES',
}


# ─── PARSER PRINCIPAL ────────────────────────────────────────────────────────

def parse_excel_bytes(file_bytes: bytes, curso: int, cuatrimestre: str, grupo: int = 1) -> list:
    """
    Parsea un archivo Excel (en bytes) y devuelve la lista de clases
    del cuatrimestre indicado.

    Returns:
        list of dict con claves:
            semana, dia, franja_label, asig_codigo, asig_nombre,
            tipo (''|'LAB'|'INFO'), subgrupo, aula_override, curso, cuatrimestre
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "Se necesita openpyxl. Instala con: pip install openpyxl --break-system-packages"
        )

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Buscar la hoja correcta
    # Formato canónico: '1C_grupo1', '2C_grupo1'
    # Formato UPCT alternativo: '2ºC 1ºQ', '2ºC 2ºQ', '1ºC 1ºQ', etc.
    cuatQ = '1' if cuatrimestre.startswith('1') else '2'
    target = f"{cuatrimestre}_grupo{grupo}"
    ws = None
    if target in wb.sheetnames:
        ws = wb[target]
    else:
        # Buscar hoja por indicador de cuatrimestre: '1ºQ', '1Q', '2ºQ', '2Q'
        for name in wb.sheetnames:
            nu = name.upper().replace('º', '')
            if f'{cuatQ}Q' in nu:
                ws = wb[name]
                break
        if ws is None:
            # Fallback posicional: 1C→primera hoja, 2C→segunda hoja (si existe)
            if cuatrimestre.startswith('2') and len(wb.sheetnames) > 1:
                ws = wb[wb.sheetnames[1]]
            elif wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
    if ws is None:
        return []

    # Leer toda la hoja en una matriz Python
    data = [list(row) for row in ws.iter_rows(values_only=True)]
    if len(data) < 5:
        return []

    # ── Localizar columnas de inicio de cada semana (fila 0) ──────────────────
    semana_cols = []   # [(num_semana, col_inicio), …]
    for ci, val in enumerate(data[0]):
        if isinstance(val, str):
            m = _SEMANA_RE.search(val)
            if m:
                semana_cols.append((int(m.group(1)), ci))

    if not semana_cols:
        return []

    # ── Localizar filas de franjas horarias (columna 3) ──────────────────────
    franja_rows = []   # [(row_idx, franja_label), …]
    for ri, row in enumerate(data):
        val = row[3] if len(row) > 3 else None
        if isinstance(val, str) and _FRANJA_RE.match(val.strip()):
            franja_rows.append((ri, val.strip()))

    # ── Parsear semana a semana ───────────────────────────────────────────────
    clases = []
    row3   = data[3] if len(data) > 3 else []

    for idx, (sem_num, sem_col) in enumerate(semana_cols):
        next_sem_col = semana_cols[idx + 1][1] if idx + 1 < len(semana_cols) else len(row3) + 100

        # Detectar columnas de cada día en la fila 3
        dia_cols = {}   # 'LUNES'→ci, 'MARTES'→ci, …
        for ci in range(sem_col + 1, next_sem_col):
            if ci >= len(row3):
                break
            val = row3[ci]
            if isinstance(val, str):
                v = val.strip().upper()
                dia_norm = DIAS_MAP.get(v)
                if dia_norm and dia_norm not in dia_cols:
                    dia_cols[dia_norm] = ci

        # Para cada franja y cada día leer el contenido
        for ri, franja_label in franja_rows:
            if ri >= len(data):
                continue
            row = data[ri]
            for dia, ci in dia_cols.items():
                if ci >= len(row):
                    continue
                val = row[ci]
                if not isinstance(val, str) or not val.strip():
                    continue
                val = val.strip()
                if 'NO LECTIVO' in val.upper():
                    continue

                parsed_list = _parse_celda(val)
                for parsed in parsed_list:
                    parsed.update({
                        'semana':        sem_num,
                        'dia':           dia,
                        'franja_label':  franja_label,
                        'curso':         curso,
                        'cuatrimestre':  cuatrimestre,
                    })
                    clases.append(parsed)

    return clases


def parse_excel_all_cuats(file_bytes: bytes, curso: int) -> dict:
    """
    Parsea los dos cuatrimestres de un archivo Excel de curso.

    Returns:
        dict: {'1C': [...], '2C': [...], 'error': str|None,
               'asignaturas': [{'codigo','nombre','curso','cuatrimestre'}, …]}
    """
    result = {'1C': [], '2C': [], 'error': None, 'asignaturas': []}
    try:
        result['1C'] = parse_excel_bytes(file_bytes, curso, '1C')
        result['2C'] = parse_excel_bytes(file_bytes, curso, '2C')
    except Exception as e:
        result['error'] = str(e)
        return result

    # Asignaturas únicas detectadas
    asig_dict = {}
    for cuat in ('1C', '2C'):
        for c in result[cuat]:
            cod = c['asig_codigo']
            if cod not in asig_dict:
                asig_dict[cod] = {
                    'codigo':        cod,
                    'nombre':        c['asig_nombre'],
                    'curso':         curso,
                    'cuatrimestre':  c['cuatrimestre'],
                }
    result['asignaturas'] = list(asig_dict.values())
    return result


# ─── PARSER DE CELDA ─────────────────────────────────────────────────────────

def _parse_celda(val: str) -> list:
    """
    Parsea el contenido de una celda de clase.

    Soporta celdas con una sola clase o varias separadas por ' / ':
        [524101003] Expresión Gráfica
        [524101001] Matemáticas I | LAB | Subgrupos: 1 | Obs: / [524101003] Expresión Gráfica | Subgrupos: 2

    Devuelve siempre una lista (vacía si la celda no contiene clases válidas).
    """
    resultados = []
    # Separar entradas múltiples: ' / [' es el delimitador de desdoble
    partes = re.split(r'\s*/\s*(?=\[)', val.strip())
    for parte in partes:
        parte = parte.strip()
        m = _SUBJECT_RE.match(parte)
        if not m:
            continue

        codigo  = m.group(1).strip()
        nombre  = m.group(2).strip()
        resto   = m.group(3) or ''

        tipo          = ''
        subgrupo      = ''
        aula_override = ''

        for part in [p.strip() for p in resto.split('|') if p.strip()]:
            pu = part.upper()
            if pu == 'LAB':
                tipo = 'LAB'
            elif pu in ('INFO', 'INF'):
                tipo = 'INF'
            elif pu.startswith('SUBGRUPOS:'):
                subgrupo = part.split(':', 1)[1].strip()
            elif pu.startswith('AULA:'):
                val_aula = part.split(':', 1)[1].strip()
                # 'Aula: LAB' / 'Aula: INF' indican tipo de actividad, no aula física
                if val_aula.upper() == 'LAB':
                    tipo = 'LAB'
                elif val_aula.upper() in ('INF', 'INFO'):
                    tipo = 'INF'
                else:
                    aula_override = val_aula
            # 'OBS:' y otros campos se ignoran intencionadamente

        resultados.append({
            'asig_codigo':    codigo,
            'asig_nombre':    nombre,
            'tipo':           tipo,
            'subgrupo':       subgrupo,
            'aula_override':  aula_override,
        })
    return resultados


# ─── CLI ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import sys, json

    if len(sys.argv) < 3:
        print("Uso: python3 importar_horarios.py <archivo.xlsx> <curso>")
        sys.exit(1)

    path   = sys.argv[1]
    curso  = int(sys.argv[2])

    with open(path, 'rb') as f:
        data = f.read()

    result = parse_excel_all_cuats(data, curso)

    print(f"1C: {len(result['1C'])} clases")
    print(f"2C: {len(result['2C'])} clases")
    print(f"Asignaturas: {len(result['asignaturas'])}")
    for a in result['asignaturas']:
        print(f"  [{a['codigo']}] {a['nombre']} (C{a['curso']} {a['cuatrimestre']})")
    if result['error']:
        print(f"Error: {result['error']}")
