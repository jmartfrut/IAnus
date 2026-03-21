#!/usr/bin/env python3
"""
exportar_excel.py — Exporta todos los horarios GIM al formato PLANTILLA_EXPORTACION.xlsx
Uso standalone: python3 exportar_excel.py [ruta_salida.xlsx]
Uso desde servidor: import exportar_excel; exportar_excel.exportar(db_path, template_path, output_path)
"""
import os
import sys
import sqlite3
import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ─── CONSTANTES DE LAYOUT (basadas en PLANTILLA_EXPORTACION.xlsx) ───

# Fila Excel donde aparece cada franja (franja_orden → fila Excel)
FRANJA_ROW = {1: 5, 2: 7, 3: 9, 4: 16, 5: 20, 6: 24}

# Offset de columna dentro de cada bloque de semana (0=etiqueta franja, 1=LUN..5=VIE, 6=separador)
DAY_OFFSET = {"LUNES": 1, "MARTES": 2, "MIÉRCOLES": 3, "JUEVES": 4, "VIERNES": 5}

# Etiquetas franja normalizadas
FRANJA_LABELS = {
    1: "9:00 - 10:50",
    2: "11:10 - 13:00",
    3: "13:10 - 15:00",
    4: "15:00 - 16:50",
    5: "17:10 - 19:00",
    6: "19:10 - 21:00",
}

# Rangos de filas para merges de franja (fila_inicio, fila_fin)
FRANJA_MERGE_ROWS = {1: (5, 6), 2: (7, 8), 3: (9, 14), 4: (16, 19), 5: (20, 23), 6: (24, 27)}

# Número de semanas
N_SEMANAS = 16

# Col Excel (1-indexed) donde empieza cada semana: 4, 11, 18, 25, ...
def week_start_col(semana_num):
    return 4 + (semana_num - 1) * 7

def day_col(semana_num, dia):
    return week_start_col(semana_num) + DAY_OFFSET[dia]

def franja_label_col(semana_num):
    return week_start_col(semana_num)


# ─── DATOS ───

def get_db(db_path):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

def fetch_all_data(conn):
    cur = conn.cursor()

    cur.execute("SELECT * FROM grupos ORDER BY curso, cuatrimestre, grupo")
    grupos = [dict(r) for r in cur.fetchall()]

    cur.execute("SELECT * FROM franjas ORDER BY orden")
    franjas = {r["orden"]: dict(r) for r in cur.fetchall()}

    data = {}
    for g in grupos:
        clave = g["clave"]
        cur.execute("""
            SELECT s.numero, s.descripcion,
                   c.dia, f.orden as franja_orden,
                   a.codigo, a.nombre,
                   c.aula, c.subgrupo, c.observacion, c.es_no_lectivo
            FROM semanas s
            JOIN grupos gr ON s.grupo_id = gr.id
            LEFT JOIN clases c ON c.semana_id = s.id
            LEFT JOIN asignaturas a ON c.asignatura_id = a.id
            LEFT JOIN franjas f ON c.franja_id = f.id
            WHERE gr.clave = ?
            ORDER BY s.numero, c.dia, f.orden
        """, (clave,))
        rows = cur.fetchall()

        # Semana descriptions
        cur.execute("""
            SELECT s.numero, s.descripcion FROM semanas s
            JOIN grupos gr ON s.grupo_id = gr.id WHERE gr.clave = ?
            ORDER BY s.numero
        """, (clave,))
        semanas_desc = {r["numero"]: r["descripcion"] for r in cur.fetchall()}

        # Organizar: {semana_num: {dia: {franja_orden: [textos]}}}
        schedule = {n: {} for n in range(1, N_SEMANAS + 1)}
        no_lectivo_dias = {n: set() for n in range(1, N_SEMANAS + 1)}

        for row in rows:
            sem = row["numero"]
            dia = row["dia"]
            fr = row["franja_orden"]
            if dia is None or fr is None:
                continue
            if row["es_no_lectivo"]:
                no_lectivo_dias[sem].add(dia)
                continue
            if row["codigo"] is None:
                continue
            texto = f"[{row['codigo']}] {row['nombre']}"
            if row["aula"]:
                texto += f" | Aula: {row['aula']}"
            if row["subgrupo"]:
                texto += f" | Subgrupos: {row['subgrupo']}"
            if row["observacion"]:
                texto += f" | Obs: {row['observacion']}"
            schedule[sem].setdefault(dia, {}).setdefault(fr, []).append(texto)

        data[clave] = {
            "grupo": g,
            "semanas_desc": semanas_desc,
            "schedule": schedule,
            "no_lectivo": no_lectivo_dias,
        }

    return grupos, data


# ─── FORMATEO ───

def make_border(style="thin"):
    s = Side(style=style, color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

BORDER_THIN = make_border("thin")
BORDER_MEDIUM = make_border("medium")

def style_header_semana(cell, text):
    cell.value = text
    cell.font = Font(bold=True, size=10, name="Arial")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER_THIN

def style_franja_label(cell, text):
    cell.value = text
    cell.font = Font(bold=True, size=9, name="Arial")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER_THIN

def style_day_header(cell, text):
    cell.value = text
    cell.font = Font(bold=True, size=9, name="Arial")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = make_fill("D9E1F2")
    cell.border = BORDER_THIN

def style_content(cell, text, no_lectivo=False):
    cell.value = text
    cell.font = Font(bold=no_lectivo, size=8, name="Arial",
                     color="CC0000" if no_lectivo else "000000")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER_THIN

def style_meta_label(cell, text):
    cell.value = text
    cell.font = Font(bold=True, size=9, name="Arial")
    cell.border = BORDER_THIN

def style_meta_value(cell, text):
    cell.value = text
    cell.font = Font(size=9, name="Arial")
    cell.border = BORDER_THIN

def style_separator(cell):
    cell.fill = make_fill("F2F2F2")
    cell.border = BORDER_THIN


# ─── CONSTRUCCIÓN DE HOJA ───

def build_sheet(ws, grupo_data, grupos_order, degree_acronym="GIM"):
    g = grupo_data["grupo"]
    semanas_desc = grupo_data["semanas_desc"]
    schedule = grupo_data["schedule"]
    no_lectivo = grupo_data["no_lectivo"]

    curso = g["curso"]
    cuat = g["cuatrimestre"]
    grupo_num = g["grupo"]
    aula = g.get("aula", "")
    clave = g["clave"]

    label_grupo = "Grupo Único" if grupo_num == "unico" else f"Grupo {grupo_num}"
    label_cuat = "1er Cuatrimestre" if cuat == "1C" else "2º Cuatrimestre"
    titulo_bloque = f"{curso}º Curso {label_cuat} ({label_grupo})"

    # ── Dimensiones de filas y columnas ──
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 19
    ws.row_dimensions[4].height = 22
    for r in [5, 6]: ws.row_dimensions[r].height = 45
    for r in [7, 8]: ws.row_dimensions[r].height = 45
    for r in range(9, 16): ws.row_dimensions[r].height = 15
    for r in range(16, 20): ws.row_dimensions[r].height = 25
    for r in range(20, 24): ws.row_dimensions[r].height = 25
    for r in range(24, 29): ws.row_dimensions[r].height = 25

    # Columnas fijas A-C
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 3

    # Columnas por semana
    for sem in range(1, N_SEMANAS + 1):
        base = week_start_col(sem)
        # Franja label col
        ws.column_dimensions[get_column_letter(base)].width = 13
        # LUNES (especial: más ancho)
        ws.column_dimensions[get_column_letter(base + 1)].width = 22
        # MARTES-VIERNES
        for d in range(2, 6):
            ws.column_dimensions[get_column_letter(base + d)].width = 18
        # Separador
        ws.column_dimensions[get_column_letter(base + 6)].width = 2

    # ── Columnas A-B: etiquetas meta ──
    meta_labels = ["Titulación:", "Curso:", "Cuatrimestre:", "Grupo:", "Aula:"]
    meta_values = [degree_acronym, curso, cuat, grupo_num if grupo_num != "unico" else "Único", aula]
    for i, (lbl, val) in enumerate(zip(meta_labels, meta_values)):
        row = i + 1
        if row > 5:
            break
        style_meta_label(ws.cell(row=row, column=1), lbl)
        style_meta_value(ws.cell(row=row, column=2), val)

    # Celdas meta vacías filas 6-28 col A-B
    for row in range(6, 29):
        for col in [1, 2]:
            ws.cell(row=row, column=col).border = BORDER_THIN

    # ── Col C: separador vertical ──
    ws.merge_cells(start_row=1, start_column=3, end_row=28, end_column=3)
    sep = ws.cell(row=1, column=3)
    sep.fill = make_fill("BFBFBF")

    # ── Por semana ──
    for sem in range(1, N_SEMANAS + 1):
        base = week_start_col(sem)
        desc = semanas_desc.get(sem, f"SEMANA {sem}")

        # ── Fila 1: cabecera SEMANA (merge D-I, 6 cols) ──
        ws.merge_cells(start_row=1, start_column=base, end_row=1, end_column=base + 5)
        style_header_semana(ws.cell(row=1, column=base), desc)

        # ── Fila 2: título bloque (merge D-I) ──
        ws.merge_cells(start_row=2, start_column=base, end_row=2, end_column=base + 5)
        c2 = ws.cell(row=2, column=base)
        c2.value = titulo_bloque
        c2.font = Font(bold=True, size=9, name="Arial")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.fill = make_fill("D9E1F2")
        c2.border = BORDER_THIN

        # ── Fila 3: vacía (separador) ──
        ws.merge_cells(start_row=3, start_column=base, end_row=3, end_column=base + 5)
        ws.cell(row=3, column=base).border = BORDER_THIN

        # ── Fila 4: cabeceras días ──
        ws.merge_cells(start_row=4, start_column=base, end_row=4, end_column=base)  # col franja
        style_franja_label(ws.cell(row=4, column=base), "Franja")
        for dia, off in DAY_OFFSET.items():
            style_day_header(ws.cell(row=4, column=base + off), dia)

        # ── Col separador (offset 6) ──
        for row in range(1, 29):
            c = ws.cell(row=row, column=base + 6)
            c.fill = make_fill("BFBFBF")

        # Determinar qué días son no lectivos en esta semana
        no_lect_dias = no_lectivo.get(sem, set())

        # ── Columnas de días: merges y contenido ──
        for dia, off in DAY_OFFSET.items():
            col = base + off
            if dia in no_lect_dias:
                # Merge toda la zona de franjas → un único "NO LECTIVO"
                ws.merge_cells(start_row=5, start_column=col, end_row=27, end_column=col)
                style_content(ws.cell(row=5, column=col), "NO LECTIVO", no_lectivo=True)
            else:
                # Merge por franja
                for fr_ord, (r1, r2) in FRANJA_MERGE_ROWS.items():
                    if r1 != r2:
                        ws.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)
                    clases = schedule.get(sem, {}).get(dia, {}).get(fr_ord, [])
                    texto = " / ".join(clases) if clases else ""
                    style_content(ws.cell(row=r1, column=col), texto)

        # ── Etiquetas de franja (col base) ──
        for fr_ord, (r1, r2) in FRANJA_MERGE_ROWS.items():
            if r1 != r2:
                ws.merge_cells(start_row=r1, start_column=base, end_row=r2, end_column=base)
            style_franja_label(ws.cell(row=r1, column=base), FRANJA_LABELS[fr_ord])

        # ── Fila 28: borde inferior ──
        ws.merge_cells(start_row=28, start_column=base, end_row=28, end_column=base + 5)
        ws.cell(row=28, column=base).border = BORDER_MEDIUM

    # ── Print area y ajuste de página ──
    ws.print_area = f"A1:{get_column_letter(week_start_col(N_SEMANAS) + 5)}28"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1


# ─── EXPORTACIÓN PRINCIPAL ───

def _sheet_name(g):
    """Genera el nombre de hoja para un grupo."""
    cuat_label = "1C" if g["cuatrimestre"] == "1C" else "2C"
    if g["grupo"] == "unico":
        return f"{cuat_label}_GrupoUnico"
    return f"{cuat_label}_Grupo{g['grupo']}"


def exportar_curso(db_path, template_path, output_path, curso_num, degree_acronym="GIM"):
    """Genera un Excel con todas las hojas de un curso (ambos cuatrimestres y grupos)."""
    conn = get_db(db_path)
    grupos, all_data = fetch_all_data(conn)
    conn.close()

    from openpyxl import Workbook
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    # Ordenar: primero 1C, luego 2C; dentro de cada cuat por número de grupo
    grupos_curso = [g for g in grupos if g["curso"] == curso_num]
    grupos_curso.sort(key=lambda g: (g["cuatrimestre"], g["grupo"]))

    for g in grupos_curso:
        ws = wb_out.create_sheet(title=_sheet_name(g))
        build_sheet(ws, all_data[g["clave"]], grupos, degree_acronym=degree_acronym)

    wb_out.save(output_path)
    return output_path


def exportar_todos_por_curso(db_path, template_path, output_dir, degree_acronym="GIM"):
    """
    Genera un Excel por curso en output_dir.
    Devuelve lista de (curso_num, ruta_archivo).
    """
    conn = get_db(db_path)
    grupos, all_data = fetch_all_data(conn)
    conn.close()

    cursos = sorted(set(g["curso"] for g in grupos))
    archivos = []

    for curso_num in cursos:
        from openpyxl import Workbook
        wb_out = Workbook()
        wb_out.remove(wb_out.active)

        grupos_curso = [g for g in grupos if g["curso"] == curso_num]
        grupos_curso.sort(key=lambda g: (g["cuatrimestre"], g["grupo"]))

        for g in grupos_curso:
            ws = wb_out.create_sheet(title=_sheet_name(g))
            build_sheet(ws, all_data[g["clave"]], grupos, degree_acronym=degree_acronym)

        path = os.path.join(output_dir, f"Horarios_{curso_num}o_curso.xlsx")
        wb_out.save(path)
        archivos.append((curso_num, path))

    return archivos


def exportar(db_path, template_path, output_path):
    """Compatibilidad: genera un único Excel con todos los grupos en hojas separadas."""
    conn = get_db(db_path)
    grupos, all_data = fetch_all_data(conn)
    conn.close()

    from openpyxl import Workbook
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for g in grupos:
        clave = g["clave"]
        curso = g["curso"]
        grupo_num = g["grupo"]
        if grupo_num == "unico":
            sheet_name = f"{curso}_{g['cuatrimestre']}_Gunico"
        else:
            sheet_name = f"{curso}_{g['cuatrimestre']}_G{grupo_num}"
        ws = wb_out.create_sheet(title=sheet_name)
        build_sheet(ws, all_data[clave], grupos)

    wb_out.save(output_path)
    return output_path


# ─── EJECUCIÓN STANDALONE ───

if __name__ == "__main__":
    db_path = os.environ.get("DB_PATH_OVERRIDE") or os.path.join(SCRIPT_DIR, "horarios.db")
    if not os.path.exists(db_path):
        # Fallback a horarios_2627.db si existe
        alt = os.path.join(SCRIPT_DIR, "horarios_2627.db")
        if os.path.exists(alt):
            db_path = alt
        else:
            print(f"ERROR: No se encuentra la BD en {db_path}")
            sys.exit(1)

    template_path = os.path.join(SCRIPT_DIR, "PLANTILLA_EXPORTACION.xlsx")
    if not os.path.exists(template_path):
        print(f"ERROR: No se encuentra {template_path}")
        sys.exit(1)

    output_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(SCRIPT_DIR, "horarios_exportados.xlsx")

    print(f"BD:       {db_path}")
    print(f"Plantilla: {template_path}")
    print(f"Salida:    {output_path}")
    print("Generando...")

    exportar(db_path, template_path, output_path)
    print(f"✓ Exportado: {output_path}")
